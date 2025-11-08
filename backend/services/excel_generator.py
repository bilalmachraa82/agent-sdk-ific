"""
ExcelGenerator - PT2030-Compliant EVF Excel Report Generation

Generates professional, multi-sheet Excel reports for EVF proposals using openpyxl.
Includes financial projections, charts, compliance checklists, and PT2030 formatting.

Key Features:
- 7 comprehensive sheets (Executive Summary, Company Info, Projections, etc.)
- Professional PT2030 color scheme (blue/green)
- Interactive charts (Revenue/Costs, Cash Flow, VALF sensitivity)
- Conditional formatting for compliance indicators
- Multi-language support (PT-PT, EN)
- Audit trail and metadata

CRITICAL: All data is deterministic - pulled from FinancialAgent and ComplianceAgent.
NO AI/LLM is used for report generation - only data formatting and presentation.
"""

from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from uuid import UUID
import io
import hashlib
import json

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
import structlog

from agents.financial_agent import FinancialOutput, CashFlow
from agents.compliance_agent import ComplianceResult, CompanyInfo
from agents.narrative_agent import NarrativeOutput
from agents.input_agent import CompanyInfo as SAFTCompanyInfo
from services.file_storage import FileStorageService
from models.file import File

logger = structlog.get_logger(__name__)


# ============================================================================
# PT2030 COLOR SCHEME & STYLES
# ============================================================================

PT2030_COLORS = {
    "primary_blue": "1E3A8A",      # Dark blue for headers
    "secondary_blue": "3B82F6",    # Medium blue for subheaders
    "accent_green": "10B981",      # Green for positive metrics
    "warning_yellow": "F59E0B",    # Yellow for warnings
    "error_red": "EF4444",         # Red for errors
    "light_gray": "F3F4F6",        # Light gray for alternating rows
    "white": "FFFFFF",
    "text_dark": "1F2937",
}

TRANSLATIONS = {
    "pt": {
        "executive_summary": "Sumário Executivo",
        "company_information": "Informação da Empresa",
        "financial_projections": "Projeções Financeiras",
        "cash_flow_analysis": "Análise de Cash Flow",
        "financial_ratios": "Rácios Financeiros",
        "compliance_checklist": "Checklist de Conformidade",
        "assumptions": "Pressupostos e Metodologia",
        "year": "Ano",
        "revenue": "Receitas",
        "costs": "Custos",
        "ebitda": "EBITDA",
        "free_cash_flow": "Free Cash Flow",
        "valf": "VALF (Valor Atual Líquido Financeiro)",
        "trf": "TRF (Taxa de Rentabilidade Financeira)",
        "compliant": "Conforme",
        "non_compliant": "Não Conforme",
        "project_name": "Nome do Projeto",
        "nif": "NIF",
        "company_name": "Nome da Empresa",
        "total_investment": "Investimento Total",
        "funding_requested": "Financiamento Solicitado",
        "generated_on": "Gerado em",
        "page": "Página",
    },
    "en": {
        "executive_summary": "Executive Summary",
        "company_information": "Company Information",
        "financial_projections": "Financial Projections (10 Years)",
        "cash_flow_analysis": "Cash Flow Analysis",
        "financial_ratios": "Financial Ratios",
        "compliance_checklist": "Compliance Checklist",
        "assumptions": "Assumptions & Methodology",
        "year": "Year",
        "revenue": "Revenue",
        "costs": "Costs",
        "ebitda": "EBITDA",
        "free_cash_flow": "Free Cash Flow",
        "valf": "VALF (Financial NPV)",
        "trf": "TRF (Financial IRR %)",
        "compliant": "Compliant",
        "non_compliant": "Non-Compliant",
        "project_name": "Project Name",
        "nif": "Tax ID",
        "company_name": "Company Name",
        "total_investment": "Total Investment",
        "funding_requested": "Funding Requested",
        "generated_on": "Generated on",
        "page": "Page",
    }
}


# ============================================================================
# PYDANTIC MODELS
# ============================================================================

class FileMetadata(BaseModel):
    """Metadata for generated Excel file."""
    id: UUID
    tenant_id: UUID
    file_name: str
    file_size_bytes: int
    sha256_hash: str
    storage_path: str
    created_at: datetime = Field(default_factory=datetime.utcnow)


class ExcelGenerationInput(BaseModel):
    """Complete input for Excel report generation."""
    project_id: UUID
    financial_output: FinancialOutput
    compliance_output: ComplianceResult
    narrative_output: NarrativeOutput
    company_info: SAFTCompanyInfo
    cash_flows: List[CashFlow]
    language: str = Field(default="pt", description="Report language (pt or en)")


# ============================================================================
# EXCEL GENERATOR SERVICE
# ============================================================================

class ExcelGenerator:
    """
    Professional Excel report generator for PT2030 EVF proposals.

    Generates multi-sheet Excel workbooks with:
    - Executive summary with key metrics
    - Company information and background
    - 10-year financial projections
    - Cash flow analysis with charts
    - Financial ratios dashboard
    - PT2030 compliance checklist
    - Assumptions and methodology
    """

    VERSION = "1.0.0"

    def __init__(self, file_storage: Optional[FileStorageService] = None):
        """
        Initialize ExcelGenerator.

        Args:
            file_storage: Optional FileStorageService for saving files
        """
        self.file_storage = file_storage

    async def generate_evf_report(
        self,
        project_id: UUID,
        financial_output: FinancialOutput,
        compliance_output: ComplianceResult,
        narrative_output: NarrativeOutput,
        company_info: SAFTCompanyInfo,
        cash_flows: List[CashFlow],
        language: str = "pt"
    ) -> bytes:
        """
        Generate complete EVF Excel report.

        Args:
            project_id: Unique project identifier
            financial_output: Results from FinancialAgent
            compliance_output: Results from ComplianceAgent
            narrative_output: Results from NarrativeAgent
            company_info: Company information from SAF-T
            cash_flows: List of annual cash flows
            language: Report language ("pt" or "en")

        Returns:
            Excel file as bytes

        Raises:
            ValueError: If input data is invalid
        """
        logger.info(
            "generating_evf_excel_report",
            project_id=str(project_id),
            language=language,
            cash_flow_years=len(cash_flows)
        )

        # Validate language
        if language not in ["pt", "en"]:
            raise ValueError(f"Unsupported language: {language}. Use 'pt' or 'en'")

        # Get translations
        t = TRANSLATIONS[language]

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define named styles
        self._create_named_styles(wb)

        # Generate all sheets
        self._create_executive_summary_sheet(
            wb, t, financial_output, compliance_output, narrative_output, company_info
        )
        self._create_company_info_sheet(
            wb, t, company_info, compliance_output
        )
        self._create_financial_projections_sheet(
            wb, t, cash_flows, financial_output
        )
        self._create_cash_flow_analysis_sheet(
            wb, t, cash_flows
        )
        self._create_financial_ratios_sheet(
            wb, t, financial_output
        )
        self._create_compliance_checklist_sheet(
            wb, t, compliance_output
        )
        self._create_assumptions_sheet(
            wb, t, financial_output, narrative_output
        )

        # Add metadata
        self._add_workbook_metadata(wb, project_id, company_info)

        # Convert to bytes
        excel_bytes = self._workbook_to_bytes(wb)

        logger.info(
            "evf_excel_report_generated",
            project_id=str(project_id),
            file_size=len(excel_bytes),
            sheets=len(wb.sheetnames)
        )

        return excel_bytes

    async def save_to_storage(
        self,
        excel_bytes: bytes,
        project_id: UUID,
        tenant_id: UUID,
        filename: str
    ) -> FileMetadata:
        """
        Save Excel file to storage backend.

        Args:
            excel_bytes: Excel file content
            project_id: Project identifier
            tenant_id: Tenant identifier
            filename: Desired filename

        Returns:
            FileMetadata with storage details

        Raises:
            RuntimeError: If file_storage not configured
            Exception: If save operation fails
        """
        if not self.file_storage:
            raise RuntimeError("FileStorageService not configured")

        logger.info(
            "saving_excel_to_storage",
            project_id=str(project_id),
            tenant_id=str(tenant_id),
            filename=filename,
            size=len(excel_bytes)
        )

        # Save file using FileStorageService
        file_obj = await self.file_storage.save_file(
            tenant_id=tenant_id,
            file_name=filename,
            file_content=excel_bytes,
            file_type="excel",
            user_id=None,  # System-generated
            metadata={
                "project_id": str(project_id),
                "generated_by": "ExcelGenerator",
                "version": self.VERSION,
                "generation_timestamp": datetime.utcnow().isoformat(),
            }
        )

        # Convert to FileMetadata
        metadata = FileMetadata(
            id=file_obj.id,
            tenant_id=file_obj.tenant_id,
            file_name=file_obj.file_name,
            file_size_bytes=file_obj.file_size_bytes,
            sha256_hash=file_obj.sha256_hash,
            storage_path=file_obj.storage_path,
            created_at=file_obj.upload_timestamp
        )

        logger.info(
            "excel_saved_to_storage",
            file_id=str(metadata.id),
            storage_path=metadata.storage_path
        )

        return metadata

    # ========================================================================
    # PRIVATE METHODS - STYLE CREATION
    # ========================================================================

    def _create_named_styles(self, wb: Workbook) -> None:
        """Create named styles for consistent formatting."""

        # Header style (dark blue background, white text)
        header_style = NamedStyle(name="header")
        header_style.font = Font(
            name="Calibri",
            size=14,
            bold=True,
            color=PT2030_COLORS["white"]
        )
        header_style.fill = PatternFill(
            start_color=PT2030_COLORS["primary_blue"],
            end_color=PT2030_COLORS["primary_blue"],
            fill_type="solid"
        )
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            bottom=Side(style="medium", color=PT2030_COLORS["primary_blue"])
        )
        wb.add_named_style(header_style)

        # Subheader style (medium blue)
        subheader_style = NamedStyle(name="subheader")
        subheader_style.font = Font(name="Calibri", size=12, bold=True)
        subheader_style.fill = PatternFill(
            start_color=PT2030_COLORS["secondary_blue"],
            end_color=PT2030_COLORS["secondary_blue"],
            fill_type="solid"
        )
        subheader_style.alignment = Alignment(horizontal="left", vertical="center")
        wb.add_named_style(subheader_style)

        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '#,##0.00 €'
        currency_style.alignment = Alignment(horizontal="right")
        wb.add_named_style(currency_style)

        # Percentage style
        percent_style = NamedStyle(name="percent")
        percent_style.number_format = '0.00%'
        percent_style.alignment = Alignment(horizontal="right")
        wb.add_named_style(percent_style)

    # ========================================================================
    # PRIVATE METHODS - SHEET CREATION
    # ========================================================================

    def _create_executive_summary_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        financial: FinancialOutput,
        compliance: ComplianceResult,
        narrative: NarrativeOutput,
        company: SAFTCompanyInfo
    ) -> None:
        """Create Executive Summary sheet with key highlights."""
        ws = wb.create_sheet(t["executive_summary"])

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

        # Title
        ws['A1'] = t["executive_summary"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 30

        # Company info
        row = 3
        ws[f'A{row}'] = t["company_name"]
        ws[f'B{row}'] = company.name
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = t["nif"]
        ws[f'B{row}'] = company.nif
        ws[f'A{row}'].font = Font(bold=True)

        # Key Financial Metrics
        row += 2
        ws[f'A{row}'] = "KEY FINANCIAL METRICS"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = t["valf"]
        ws[f'B{row}'] = float(financial.valf)
        ws[f'B{row}'].style = "currency"
        ws[f'A{row}'].font = Font(bold=True)

        # Color code VALF (green if negative - good for PT2030)
        if financial.valf < 0:
            ws[f'B{row}'].fill = PatternFill(
                start_color=PT2030_COLORS["accent_green"],
                fill_type="solid"
            )

        row += 1
        ws[f'A{row}'] = t["trf"]
        ws[f'B{row}'] = float(financial.trf) / 100  # Convert to decimal for %
        ws[f'B{row}'].style = "percent"
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = t["total_investment"]
        # Get total investment from compliance output
        ws[f'B{row}'] = float(compliance.calculated_funding_amount)
        ws[f'B{row}'].style = "currency"
        ws[f'A{row}'].font = Font(bold=True)

        # Compliance Status
        row += 2
        ws[f'A{row}'] = "PT2030 COMPLIANCE STATUS"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = "Overall Status"
        ws[f'B{row}'] = t["compliant"] if compliance.is_compliant else t["non_compliant"]
        ws[f'A{row}'].font = Font(bold=True)

        # Color code compliance
        if compliance.is_compliant:
            ws[f'B{row}'].fill = PatternFill(
                start_color=PT2030_COLORS["accent_green"],
                fill_type="solid"
            )
        else:
            ws[f'B{row}'].fill = PatternFill(
                start_color=PT2030_COLORS["error_red"],
                fill_type="solid"
            )

        row += 1
        ws[f'A{row}'] = "Critical Issues"
        ws[f'B{row}'] = compliance.critical_failures
        ws[f'A{row}'].font = Font(bold=True)

        # Executive Summary Text
        row += 2
        ws[f'A{row}'] = "EXECUTIVE SUMMARY"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = narrative.executive_summary
        ws.merge_cells(f'A{row}:B{row+10}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")

    def _create_company_info_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        company: SAFTCompanyInfo,
        compliance: ComplianceResult
    ) -> None:
        """Create Company Information sheet."""
        ws = wb.create_sheet(t["company_information"])

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

        # Title
        ws['A1'] = t["company_information"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:B1')

        # Company details
        row = 3
        fields = [
            ("Legal Name", company.name),
            (t["nif"], company.nif),
            ("Fiscal Year", company.fiscal_year),
            ("Currency", company.currency_code),
            ("SAF-T Version", company.audit_file_version or "N/A"),
        ]

        for label, value in fields:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Address (if available)
        if company.address:
            row += 1
            ws[f'A{row}'] = "ADDRESS"
            ws[f'A{row}'].style = "subheader"
            ws.merge_cells(f'A{row}:B{row}')

            row += 1
            address_parts = []
            if company.address.address_detail:
                address_parts.append(company.address.address_detail)
            if company.address.postal_code and company.address.city:
                address_parts.append(f"{company.address.postal_code} {company.address.city}")
            if company.address.region:
                address_parts.append(company.address.region)

            ws[f'A{row}'] = "\n".join(address_parts)
            ws.merge_cells(f'A{row}:B{row+2}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)

    def _create_financial_projections_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        cash_flows: List[CashFlow],
        financial: FinancialOutput
    ) -> None:
        """Create Financial Projections sheet with 10-year data and charts."""
        ws = wb.create_sheet(t["financial_projections"])

        # Title
        ws['A1'] = t["financial_projections"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:L1')

        # Headers
        row = 3
        headers = [
            t["year"],
            t["revenue"],
            "Operating Costs",
            "CAPEX",
            "Depreciation",
            t["ebitda"],
            "EBIT",
            t["free_cash_flow"],
            "Cumulative FCF"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=PT2030_COLORS["light_gray"],
                fill_type="solid"
            )

        # Data rows
        cumulative_fcf = Decimal("0")
        for cf in cash_flows:
            row += 1
            cumulative_fcf += cf.free_cash_flow

            ws.cell(row=row, column=1).value = cf.year
            ws.cell(row=row, column=2).value = float(cf.revenue)
            ws.cell(row=row, column=2).style = "currency"
            ws.cell(row=row, column=3).value = float(cf.operating_costs)
            ws.cell(row=row, column=3).style = "currency"
            ws.cell(row=row, column=4).value = float(cf.capex)
            ws.cell(row=row, column=4).style = "currency"
            ws.cell(row=row, column=5).value = float(cf.depreciation)
            ws.cell(row=row, column=5).style = "currency"
            ws.cell(row=row, column=6).value = float(cf.ebitda)
            ws.cell(row=row, column=6).style = "currency"
            ws.cell(row=row, column=7).value = float(cf.ebit)
            ws.cell(row=row, column=7).style = "currency"
            ws.cell(row=row, column=8).value = float(cf.free_cash_flow)
            ws.cell(row=row, column=8).style = "currency"
            ws.cell(row=row, column=9).value = float(cumulative_fcf)
            ws.cell(row=row, column=9).style = "currency"

        # Add Revenue vs Costs chart
        self._add_revenue_costs_chart(ws, len(cash_flows))

        # Auto-size columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_cash_flow_analysis_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        cash_flows: List[CashFlow]
    ) -> None:
        """Create Cash Flow Analysis sheet with waterfall chart."""
        ws = wb.create_sheet(t["cash_flow_analysis"])

        ws['A1'] = t["cash_flow_analysis"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:F1')

        # Headers
        row = 3
        headers = [
            t["year"],
            t["revenue"],
            "EBITDA",
            "CAPEX",
            "Working Capital Δ",
            t["free_cash_flow"]
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=PT2030_COLORS["secondary_blue"],
                fill_type="solid"
            )

        # Data
        for cf in cash_flows:
            row += 1
            ws.cell(row=row, column=1).value = cf.year
            ws.cell(row=row, column=2).value = float(cf.revenue)
            ws.cell(row=row, column=2).style = "currency"
            ws.cell(row=row, column=3).value = float(cf.ebitda)
            ws.cell(row=row, column=3).style = "currency"
            ws.cell(row=row, column=4).value = float(cf.capex)
            ws.cell(row=row, column=4).style = "currency"
            ws.cell(row=row, column=5).value = float(cf.working_capital_change)
            ws.cell(row=row, column=5).style = "currency"
            ws.cell(row=row, column=6).value = float(cf.free_cash_flow)
            ws.cell(row=row, column=6).style = "currency"

            # Conditional formatting for FCF
            if cf.free_cash_flow >= 0:
                ws.cell(row=row, column=6).font = Font(
                    color=PT2030_COLORS["accent_green"]
                )
            else:
                ws.cell(row=row, column=6).font = Font(
                    color=PT2030_COLORS["error_red"]
                )

        # Add cash flow chart
        self._add_cash_flow_chart(ws, len(cash_flows))

    def _create_financial_ratios_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        financial: FinancialOutput
    ) -> None:
        """Create Financial Ratios dashboard."""
        ws = wb.create_sheet(t["financial_ratios"])

        ws['A1'] = t["financial_ratios"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:C1')

        row = 3
        ws[f'A{row}'] = "Ratio Category"
        ws[f'B{row}'] = "Metric"
        ws[f'C{row}'] = "Value"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(
                start_color=PT2030_COLORS["light_gray"],
                fill_type="solid"
            )

        ratios = financial.financial_ratios

        # Profitability Ratios
        row += 1
        metrics = [
            ("Profitability", "Gross Margin", ratios.gross_margin),
            ("", "Operating Margin", ratios.operating_margin),
            ("", "Net Margin", ratios.net_margin),
            ("Returns", "ROI", ratios.roi),
            ("", "ROIC", ratios.roic),
            ("Efficiency", "Asset Turnover", ratios.asset_turnover),
            ("", "CAPEX/Revenue", ratios.capex_to_revenue),
            ("Coverage", "EBITDA Coverage", ratios.ebitda_coverage),
            ("", "FCF Coverage", ratios.fcf_coverage),
        ]

        for category, metric, value in metrics:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = metric
            if value is not None:
                ws[f'C{row}'] = float(value)
                # Format as percentage if applicable
                if metric in ["Gross Margin", "Operating Margin", "Net Margin", "ROI", "ROIC"]:
                    ws[f'C{row}'].style = "percent"
            else:
                ws[f'C{row}'] = "N/A"
            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15

    def _create_compliance_checklist_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        compliance: ComplianceResult
    ) -> None:
        """Create PT2030 Compliance Checklist."""
        ws = wb.create_sheet(t["compliance_checklist"])

        ws['A1'] = t["compliance_checklist"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:E1')

        # Headers
        row = 3
        headers = ["Check ID", "Check Name", "Severity", "Status", "Message"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=PT2030_COLORS["secondary_blue"],
                fill_type="solid"
            )

        # Compliance checks
        for check in compliance.checks:
            row += 1
            ws.cell(row=row, column=1).value = check.check_id
            ws.cell(row=row, column=2).value = check.check_name
            ws.cell(row=row, column=3).value = check.severity.upper()
            ws.cell(row=row, column=4).value = "PASS" if check.passed else "FAIL"
            ws.cell(row=row, column=5).value = check.message

            # Color code status
            status_cell = ws.cell(row=row, column=4)
            if check.passed:
                status_cell.fill = PatternFill(
                    start_color=PT2030_COLORS["accent_green"],
                    fill_type="solid"
                )
            else:
                if check.severity == "critical":
                    status_cell.fill = PatternFill(
                        start_color=PT2030_COLORS["error_red"],
                        fill_type="solid"
                    )
                else:
                    status_cell.fill = PatternFill(
                        start_color=PT2030_COLORS["warning_yellow"],
                        fill_type="solid"
                    )

        # Summary
        row += 2
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:E{row}')

        row += 1
        ws[f'A{row}'] = "Overall Compliance"
        ws[f'B{row}'] = t["compliant"] if compliance.is_compliant else t["non_compliant"]

        row += 1
        ws[f'A{row}'] = "Critical Failures"
        ws[f'B{row}'] = compliance.critical_failures

        row += 1
        ws[f'A{row}'] = "Warnings"
        ws[f'B{row}'] = compliance.warnings

        # Auto-size columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 50

    def _create_assumptions_sheet(
        self,
        wb: Workbook,
        t: Dict[str, str],
        financial: FinancialOutput,
        narrative: NarrativeOutput
    ) -> None:
        """Create Assumptions & Methodology sheet."""
        ws = wb.create_sheet(t["assumptions"])

        ws['A1'] = t["assumptions"]
        ws['A1'].style = "header"
        ws.merge_cells('A1:B1')

        # Financial Assumptions
        row = 3
        ws[f'A{row}'] = "FINANCIAL ASSUMPTIONS"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        for key, value in financial.assumptions.items():
            ws[f'A{row}'] = key.replace("_", " ").title()
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Methodology
        row += 1
        ws[f'A{row}'] = "METHODOLOGY"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = narrative.methodology
        ws.merge_cells(f'A{row}:B{row+10}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")

        # Calculation Details
        row += 12
        ws[f'A{row}'] = "CALCULATION DETAILS"
        ws[f'A{row}'].style = "subheader"
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = "Calculation Timestamp"
        ws[f'B{row}'] = financial.calculation_timestamp.isoformat()
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Calculation Method"
        ws[f'B{row}'] = financial.calculation_method
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Input Data Hash"
        ws[f'B{row}'] = financial.input_hash
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(size=8)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60

    # ========================================================================
    # PRIVATE METHODS - CHART CREATION
    # ========================================================================

    def _add_revenue_costs_chart(self, ws: Worksheet, num_years: int) -> None:
        """Add Revenue vs Costs line chart."""
        chart = LineChart()
        chart.title = "Revenue vs Operating Costs"
        chart.style = 10
        chart.y_axis.title = "EUR"
        chart.x_axis.title = "Year"

        # Revenue data (column B)
        revenue_data = Reference(ws, min_col=2, min_row=3, max_row=3+num_years)
        revenue_series = Series(revenue_data, title="Revenue")
        chart.series.append(revenue_series)

        # Costs data (column C)
        costs_data = Reference(ws, min_col=3, min_row=3, max_row=3+num_years)
        costs_series = Series(costs_data, title="Operating Costs")
        chart.series.append(costs_series)

        # Year labels (column A)
        years = Reference(ws, min_col=1, min_row=4, max_row=3+num_years)
        chart.set_categories(years)

        # Position chart
        ws.add_chart(chart, "K3")

    def _add_cash_flow_chart(self, ws: Worksheet, num_years: int) -> None:
        """Add Free Cash Flow bar chart."""
        chart = BarChart()
        chart.type = "col"
        chart.title = "Free Cash Flow by Year"
        chart.style = 10
        chart.y_axis.title = "EUR"
        chart.x_axis.title = "Year"

        # FCF data (column F)
        fcf_data = Reference(ws, min_col=6, min_row=3, max_row=3+num_years)
        fcf_series = Series(fcf_data, title="Free Cash Flow")
        chart.series.append(fcf_series)

        # Year labels
        years = Reference(ws, min_col=1, min_row=4, max_row=3+num_years)
        chart.set_categories(years)

        # Position chart
        ws.add_chart(chart, "H3")

    # ========================================================================
    # PRIVATE METHODS - UTILITIES
    # ========================================================================

    def _add_workbook_metadata(
        self,
        wb: Workbook,
        project_id: UUID,
        company: SAFTCompanyInfo
    ) -> None:
        """Add metadata to workbook properties."""
        wb.properties.title = f"EVF Report - {company.name}"
        wb.properties.subject = "PT2030 Economic-Financial Viability Study"
        wb.properties.creator = "EVF Portugal 2030 Platform"
        wb.properties.description = f"Project ID: {project_id}"
        wb.properties.keywords = "PT2030, EVF, VALF, TRF, Financial Analysis"
        wb.properties.created = datetime.utcnow()

    def _workbook_to_bytes(self, wb: Workbook) -> bytes:
        """Convert workbook to bytes for download/storage."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def calculate_file_hash(excel_bytes: bytes) -> str:
        """Calculate SHA-256 hash of Excel file."""
        return hashlib.sha256(excel_bytes).hexdigest()
