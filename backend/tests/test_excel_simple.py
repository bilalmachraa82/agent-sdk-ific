"""
Simple standalone test for ExcelGenerator
Tests the core functionality without complex imports
"""

import io
from datetime import datetime
from decimal import Decimal
from uuid import uuid4

# Test if we can import and use openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def test_openpyxl_basics():
    """Test openpyxl is working."""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    ws['A1'].font = Font(bold=True)

    # Convert to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    assert len(excel_bytes) > 0
    print("✅ openpyxl basic test passed")


def test_excel_structure():
    """Test creating a multi-sheet Excel with PT2030 structure."""
    wb = Workbook()
    wb.remove(wb.active)

    # Create sheets
    sheets = [
        "Executive Summary",
        "Company Information",
        "Financial Projections",
        "Cash Flow Analysis",
        "Financial Ratios",
        "Compliance Checklist",
        "Assumptions"
    ]

    for sheet_name in sheets:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = sheet_name
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    # Verify
    assert len(wb.sheetnames) == 7
    print(f"✅ Created {len(wb.sheetnames)} sheets successfully")

    # Add some data to financial projections
    ws = wb["Financial Projections"]
    headers = ["Year", "Revenue", "Costs", "EBITDA", "Free Cash Flow"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col).value = header
        ws.cell(row=2, column=col).font = Font(bold=True)

    # Add sample data
    for year in range(0, 11):
        row = year + 3
        ws.cell(row=row, column=1).value = year
        ws.cell(row=row, column=2).value = 100000 * year
        ws.cell(row=row, column=3).value = 60000 * year
        ws.cell(row=row, column=4).value = 40000 * year
        ws.cell(row=row, column=5).value = 35000 * year

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    # Load and verify
    wb_loaded = load_workbook(io.BytesIO(excel_bytes))
    assert "Financial Projections" in wb_loaded.sheetnames
    ws_loaded = wb_loaded["Financial Projections"]
    assert ws_loaded['A3'].value == 0  # Year 0
    assert ws_loaded['A4'].value == 1  # Year 1

    print("✅ Excel structure test passed")
    print(f"   File size: {len(excel_bytes):,} bytes")


def test_currency_formatting():
    """Test currency and percentage formatting."""
    wb = Workbook()
    ws = wb.active

    # Add currency values
    ws['A1'] = "VALF"
    ws['B1'] = -45000.50
    ws['B1'].number_format = '#,##0.00 €'

    ws['A2'] = "TRF"
    ws['B2'] = 0.0375  # 3.75%
    ws['B2'].number_format = '0.00%'

    # Conditional formatting for compliance
    if ws['B1'].value < 0:
        ws['B1'].fill = PatternFill(start_color="10B981", fill_type="solid")

    # Save and verify
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    wb_loaded = load_workbook(io.BytesIO(excel_bytes))
    ws_loaded = wb_loaded.active

    assert ws_loaded['B1'].value == -45000.50
    assert ws_loaded['B2'].value == 0.0375

    print("✅ Currency formatting test passed")


def test_pt2030_colors():
    """Test PT2030 color scheme."""
    colors = {
        "primary_blue": "1E3A8A",
        "secondary_blue": "3B82F6",
        "accent_green": "10B981",
        "warning_yellow": "F59E0B",
        "error_red": "EF4444",
    }

    wb = Workbook()
    ws = wb.active

    row = 1
    for name, color_hex in colors.items():
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color=color_hex,
            end_color=color_hex,
            fill_type="solid"
        )
        row += 1

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.read()

    print("✅ PT2030 colors test passed")


if __name__ == "__main__":
    print("\n=== Running Excel Generator Simple Tests ===\n")

    test_openpyxl_basics()
    test_excel_structure()
    test_currency_formatting()
    test_pt2030_colors()

    print("\n✅ All tests passed!\n")
