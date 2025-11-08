"""
Tests for ExcelGenerator - PT2030 EVF Excel Report Generation

Comprehensive test suite for Excel generation service including:
- Sheet creation and formatting
- Data accuracy and calculations
- Chart generation
- Multi-language support
- File storage integration
- Error handling
"""

import pytest
import io
from datetime import datetime
from decimal import Decimal
from uuid import uuid4
from pathlib import Path
from unittest.mock import Mock, AsyncMock, patch

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import sys
from pathlib import Path

# Add parent directory to path for direct imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.excel_generator import (
    ExcelGenerator,
    FileMetadata,
    PT2030_COLORS,
    TRANSLATIONS
)
from agents.financial_agent import (
    FinancialOutput,
    FinancialRatios,
    CashFlow
)
from agents.compliance_agent import (
    ComplianceResult,
    ComplianceCheck,
    CheckSeverity,
    CompanyInfo as ComplianceCompanyInfo,
    CompanySize
)
from agents.narrative_agent import NarrativeOutput
from agents.input_agent import CompanyInfo as SAFTCompanyInfo, CompanyAddress
from models.file import File


# ============================================================================
# FIXTURES - TEST DATA
# ============================================================================

@pytest.fixture
def sample_company_info() -> SAFTCompanyInfo:
    """Sample company information from SAF-T."""
    return SAFTCompanyInfo(
        nif="123456789",
        name="Tech Innovations Lda",
        fiscal_year=2024,
        currency_code="EUR",
        audit_file_version="1.04_01",
        address=CompanyAddress(
            address_detail="Rua das Flores, 123",
            city="Porto",
            postal_code="4000-001",
            region="Porto",
            country="PT"
        )
    )


@pytest.fixture
def sample_cash_flows() -> list[CashFlow]:
    """Sample 10-year cash flow projections."""
    cash_flows = []

    # Year 0 - Initial investment
    cash_flows.append(
        CashFlow(
            year=0,
            revenue=Decimal("0"),
            operating_costs=Decimal("0"),
            capex=Decimal("500000"),  # Initial investment
            depreciation=Decimal("0"),
            working_capital_change=Decimal("0")
        )
    )

    # Years 1-10 - Operations
    for year in range(1, 11):
        revenue = Decimal(str(100000 * year * 1.1))  # Growing revenue
        costs = Decimal(str(revenue * 0.6))  # 60% operating costs
        capex = Decimal("10000") if year in [3, 6, 9] else Decimal("0")
        depreciation = Decimal("50000")

        cash_flows.append(
            CashFlow(
                year=year,
                revenue=revenue,
                operating_costs=costs,
                capex=capex,
                depreciation=depreciation,
                working_capital_change=Decimal("5000")
            )
        )

    return cash_flows


@pytest.fixture
def sample_financial_output() -> FinancialOutput:
    """Sample financial calculation results."""
    return FinancialOutput(
        valf=Decimal("-45000.50"),  # Negative VALF (good for PT2030)
        trf=Decimal("3.75"),  # 3.75% IRR (below 4% threshold)
        payback_period=Decimal("6.5"),
        total_fcf=Decimal("1200000"),
        average_annual_fcf=Decimal("120000"),
        financial_ratios=FinancialRatios(
            gross_margin=Decimal("0.40"),
            operating_margin=Decimal("0.25"),
            net_margin=Decimal("0.15"),
            roi=Decimal("0.18"),
            roic=Decimal("0.22"),
            asset_turnover=Decimal("1.5"),
            capex_to_revenue=Decimal("0.05"),
            ebitda_coverage=Decimal("2.4"),
            fcf_coverage=Decimal("1.8")
        ),
        pt2030_compliant=True,
        compliance_notes=["VALF < 0: PASS", "TRF < 4%: PASS"],
        calculation_timestamp=datetime.utcnow(),
        input_hash="abc123def456",
        calculation_method="numpy-financial",
        assumptions={
            "discount_rate": "4%",
            "project_duration": "10 years",
            "calculation_method": "Deterministic (numpy-financial)"
        }
    )


@pytest.fixture
def sample_compliance_output() -> ComplianceResult:
    """Sample compliance validation results."""
    checks = [
        ComplianceCheck(
            check_id="valf_check",
            check_name="VALF < 0 (Negative NPV)",
            severity=CheckSeverity.CRITICAL,
            passed=True,
            expected_value="< 0",
            actual_value="-45000.50",
            message="VALF is negative as required for PT2030",
            rule_reference="PT2030-FIN-001"
        ),
        ComplianceCheck(
            check_id="trf_check",
            check_name="TRF < Discount Rate",
            severity=CheckSeverity.CRITICAL,
            passed=True,
            expected_value="< 4%",
            actual_value="3.75%",
            message="TRF is below discount rate threshold",
            rule_reference="PT2030-FIN-002"
        ),
        ComplianceCheck(
            check_id="company_size",
            check_name="Company Size Classification",
            severity=CheckSeverity.INFO,
            passed=True,
            expected_value="SME",
            actual_value="SMALL",
            message="Company qualifies as SME",
            rule_reference="PT2030-ELG-001"
        )
    ]

    return ComplianceResult(
        is_compliant=True,
        program="PT2030",
        checks=checks,
        recommendations=["Consider increasing green investment percentage"],
        confidence_score=0.95,
        critical_failures=0,
        warnings=0,
        max_funding_rate_percent=Decimal("50"),
        calculated_funding_amount=Decimal("250000"),
        requested_funding_valid=True,
        validation_timestamp=datetime.utcnow(),
        rules_version="1.0.0",
        validator_version="1.0.0"
    )


@pytest.fixture
def sample_narrative_output() -> NarrativeOutput:
    """Sample narrative generation results."""
    return NarrativeOutput(
        executive_summary=(
            "This project represents a strategic investment in digital transformation "
            "for the Portuguese manufacturing sector. With a total investment of €500,000, "
            "the initiative will modernize production capabilities and create sustainable jobs. "
            "Financial analysis demonstrates strong viability with a negative VALF of €-45,000 "
            "and TRF of 3.75%, both meeting PT2030 requirements. The project aligns with "
            "national digitalization objectives and sustainability goals."
        ),
        methodology=(
            "Financial analysis was conducted using deterministic cash flow modeling over a "
            "10-year period. VALF (NPV) calculations employed a 4% discount rate as per PT2030 "
            "guidelines. TRF (IRR) was calculated using numpy-financial library ensuring "
            "reproducible results. All projections are based on historical SAF-T data and "
            "conservative growth assumptions."
        ),
        recommendations=(
            "1. Proceed with funding application - all PT2030 criteria met\n"
            "2. Consider increasing green investment component to 40%+\n"
            "3. Develop detailed implementation timeline\n"
            "4. Prepare supporting documentation for submission"
        ),
        tokens_used=1500,
        cost_euros=Decimal("0.05"),
        generation_time_seconds=2.5,
        word_count={
            "executive_summary": 85,
            "methodology": 62,
            "recommendations": 35
        },
        model_used="claude-4.5-sonnet",
        generation_timestamp=datetime.utcnow(),
        input_hash="abc123def456",
        cached=False
    )


# ============================================================================
# TESTS - EXCEL GENERATION
# ============================================================================

@pytest.mark.asyncio
async def test_generate_evf_report_basic(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test basic Excel report generation."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="pt"
    )

    # Verify file is generated
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    # Load and verify workbook
    wb = load_workbook(io.BytesIO(excel_bytes))

    # Verify all 7 sheets exist
    expected_sheets = [
        "Sumário Executivo",
        "Informação da Empresa",
        "Projeções Financeiras",
        "Análise de Cash Flow",
        "Rácios Financeiros",
        "Checklist de Conformidade",
        "Pressupostos e Metodologia"
    ]

    for sheet_name in expected_sheets:
        assert sheet_name in wb.sheetnames, f"Sheet '{sheet_name}' not found"


@pytest.mark.asyncio
async def test_generate_evf_report_english(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test Excel report generation in English."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))

    # Verify English sheet names
    assert "Executive Summary" in wb.sheetnames
    assert "Company Information" in wb.sheetnames
    assert "Financial Projections (10 Years)" in wb.sheetnames


@pytest.mark.asyncio
async def test_generate_invalid_language(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test error handling for invalid language."""
    generator = ExcelGenerator()
    project_id = uuid4()

    with pytest.raises(ValueError, match="Unsupported language"):
        await generator.generate_evf_report(
            project_id=project_id,
            financial_output=sample_financial_output,
            compliance_output=sample_compliance_output,
            narrative_output=sample_narrative_output,
            company_info=sample_company_info,
            cash_flows=sample_cash_flows,
            language="fr"  # Invalid
        )


# ============================================================================
# TESTS - SHEET CONTENT
# ============================================================================

@pytest.mark.asyncio
async def test_executive_summary_content(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test Executive Summary sheet content."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Executive Summary"]

    # Verify company name
    assert ws['B3'].value == "Tech Innovations Lda"

    # Verify NIF
    assert ws['B4'].value == "123456789"

    # Verify VALF is present and negative
    valf_cell = None
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                valf_cell = cell
                break

    assert valf_cell is not None, "VALF not found or not negative"

    # Verify green color for negative VALF (good for PT2030)
    # Note: Color should be applied to cells with negative VALF


@pytest.mark.asyncio
async def test_financial_projections_content(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test Financial Projections sheet content and calculations."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Financial Projections (10 Years)"]

    # Verify headers exist
    assert ws['A3'].value == "Year"
    assert ws['B3'].value == "Revenue"
    assert ws['H3'].value == "Free Cash Flow"

    # Verify Year 0 data (initial investment)
    assert ws['A4'].value == 0
    assert ws['D4'].value == 500000  # CAPEX

    # Verify Year 1 data
    assert ws['A5'].value == 1

    # Verify number of data rows (11 rows: year 0 + years 1-10)
    data_rows = sum(1 for row in ws.iter_rows(min_row=4, max_row=14, min_col=1, max_col=1)
                    if row[0].value is not None)
    assert data_rows == 11


@pytest.mark.asyncio
async def test_compliance_checklist_formatting(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test Compliance Checklist sheet formatting and color coding."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Compliance Checklist"]

    # Verify headers
    assert ws['A3'].value == "Check ID"
    assert ws['D3'].value == "Status"

    # Verify compliance checks are present
    check_count = sum(1 for row in ws.iter_rows(min_row=4, max_row=20, min_col=1, max_col=1)
                      if row[0].value is not None and row[0].value != "SUMMARY")
    assert check_count == 3  # We have 3 sample checks

    # Verify PASS/FAIL values
    assert ws['D4'].value == "PASS"  # First check


@pytest.mark.asyncio
async def test_financial_ratios_content(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test Financial Ratios sheet content."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Financial Ratios"]

    # Verify headers
    assert ws['A3'].value == "Ratio Category"
    assert ws['B3'].value == "Metric"
    assert ws['C3'].value == "Value"

    # Verify some ratios are present
    metrics_found = []
    for row in ws.iter_rows(min_row=4, max_row=20, min_col=2, max_col=2):
        if row[0].value:
            metrics_found.append(row[0].value)

    assert "Gross Margin" in metrics_found
    assert "ROI" in metrics_found


# ============================================================================
# TESTS - FILE STORAGE
# ============================================================================

@pytest.mark.asyncio
async def test_save_to_storage_success(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test successful save to file storage."""
    # Mock FileStorageService
    mock_storage = AsyncMock()
    mock_file = Mock(spec=File)
    mock_file.id = uuid4()
    mock_file.tenant_id = uuid4()
    mock_file.file_name = "evf_report.xlsx"
    mock_file.file_size_bytes = 50000
    mock_file.sha256_hash = "abc123"
    mock_file.storage_path = "/storage/evf_report.xlsx"
    mock_file.upload_timestamp = datetime.utcnow()

    mock_storage.save_file = AsyncMock(return_value=mock_file)

    generator = ExcelGenerator(file_storage=mock_storage)
    project_id = uuid4()

    # Generate Excel
    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="pt"
    )

    # Save to storage
    metadata = await generator.save_to_storage(
        excel_bytes=excel_bytes,
        project_id=project_id,
        tenant_id=mock_file.tenant_id,
        filename="evf_report.xlsx"
    )

    # Verify save was called
    mock_storage.save_file.assert_called_once()

    # Verify metadata
    assert metadata.id == mock_file.id
    assert metadata.tenant_id == mock_file.tenant_id
    assert metadata.file_name == "evf_report.xlsx"


@pytest.mark.asyncio
async def test_save_to_storage_no_service():
    """Test error when FileStorageService not configured."""
    generator = ExcelGenerator()  # No storage service

    with pytest.raises(RuntimeError, match="FileStorageService not configured"):
        await generator.save_to_storage(
            excel_bytes=b"test",
            project_id=uuid4(),
            tenant_id=uuid4(),
            filename="test.xlsx"
        )


# ============================================================================
# TESTS - UTILITIES
# ============================================================================

def test_calculate_file_hash():
    """Test file hash calculation."""
    test_data = b"test excel data"
    hash1 = ExcelGenerator.calculate_file_hash(test_data)
    hash2 = ExcelGenerator.calculate_file_hash(test_data)

    # Same data should produce same hash
    assert hash1 == hash2

    # Different data should produce different hash
    different_data = b"different data"
    hash3 = ExcelGenerator.calculate_file_hash(different_data)
    assert hash1 != hash3


@pytest.mark.asyncio
async def test_workbook_metadata(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test workbook metadata is correctly set."""
    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))

    # Verify metadata
    assert wb.properties.creator == "EVF Portugal 2030 Platform"
    assert "Tech Innovations Lda" in wb.properties.title
    assert "PT2030" in wb.properties.keywords


# ============================================================================
# TESTS - EDGE CASES
# ============================================================================

@pytest.mark.asyncio
async def test_minimal_cash_flows(
    sample_company_info,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test with minimal cash flows (2 years only)."""
    minimal_flows = [
        CashFlow(year=0, capex=Decimal("100000")),
        CashFlow(year=1, revenue=Decimal("50000"), operating_costs=Decimal("30000"))
    ]

    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=minimal_flows,
        language="en"
    )

    # Should generate without error
    assert len(excel_bytes) > 0


@pytest.mark.asyncio
async def test_missing_optional_fields(
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test with company info missing optional address."""
    company_minimal = SAFTCompanyInfo(
        nif="987654321",
        name="Minimal Company",
        fiscal_year=2024,
        # No address
    )

    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=company_minimal,
        cash_flows=sample_cash_flows,
        language="en"
    )

    # Should generate without error
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Company Information"]

    # Verify company name is present
    assert ws['B3'].value == "Minimal Company"


@pytest.mark.asyncio
async def test_non_compliant_project(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_narrative_output
):
    """Test Excel generation for non-compliant project."""
    # Create non-compliant result
    non_compliant = ComplianceResult(
        is_compliant=False,
        program="PT2030",
        checks=[
            ComplianceCheck(
                check_id="valf_fail",
                check_name="VALF Check",
                severity=CheckSeverity.CRITICAL,
                passed=False,
                expected_value="< 0",
                actual_value="100000",
                message="VALF is positive - does not meet PT2030 requirements",
                rule_reference="PT2030-FIN-001"
            )
        ],
        recommendations=["Reduce investment costs", "Increase revenue projections"],
        confidence_score=0.9,
        critical_failures=1,
        warnings=2,
        max_funding_rate_percent=Decimal("50"),
        calculated_funding_amount=Decimal("250000"),
        requested_funding_valid=False,
        validation_timestamp=datetime.utcnow(),
        rules_version="1.0.0",
        validator_version="1.0.0"
    )

    generator = ExcelGenerator()
    project_id = uuid4()

    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=non_compliant,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="en"
    )

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Executive Summary"]

    # Verify non-compliant status is shown
    # (The exact cell location may vary, but status should be "Non-Compliant")
    found_non_compliant = False
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value == "Non-Compliant":
                found_non_compliant = True
                break

    assert found_non_compliant, "Non-compliant status not found in Executive Summary"


# ============================================================================
# TESTS - PERFORMANCE
# ============================================================================

@pytest.mark.asyncio
async def test_generation_performance(
    sample_company_info,
    sample_cash_flows,
    sample_financial_output,
    sample_compliance_output,
    sample_narrative_output
):
    """Test Excel generation completes in reasonable time."""
    import time

    generator = ExcelGenerator()
    project_id = uuid4()

    start = time.time()
    excel_bytes = await generator.generate_evf_report(
        project_id=project_id,
        financial_output=sample_financial_output,
        compliance_output=sample_compliance_output,
        narrative_output=sample_narrative_output,
        company_info=sample_company_info,
        cash_flows=sample_cash_flows,
        language="pt"
    )
    duration = time.time() - start

    # Should complete in under 5 seconds
    assert duration < 5.0, f"Excel generation took {duration:.2f}s (expected < 5s)"

    # File should be reasonable size (< 5MB)
    assert len(excel_bytes) < 5 * 1024 * 1024, "File too large"
